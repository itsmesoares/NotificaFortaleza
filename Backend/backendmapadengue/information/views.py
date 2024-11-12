import json
import openpyxl
import unicodedata
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render
from django.urls import path
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from drf_yasg.views import get_schema_view
from rest_framework import permissions
import os

import importlib
importlib.invalidate_caches()


# Função para remover caracteres especiais e acentos
def limpar_nome_bairro(nome_bairro):
    if nome_bairro:
        nome_sem_acentos = ''.join((c for c in unicodedata.normalize('NFD', nome_bairro) if unicodedata.category(c) != 'Mn'))
        nome_sem_especiais = ''.join(e for e in nome_sem_acentos if e.isalnum() or e.isspace())
        return nome_sem_especiais.strip()
    return None

excel_path = r"C:\Users\isabe\OneDrive\Área de Trabalho\NotificaFortalezaTestes\Backend\backendmapadengue\information\arquivos\DengueFortaleza.xlsx"


# Verificar se o arquivo existe
if not os.path.exists(excel_path):
    raise FileNotFoundError(f"Arquivo não encontrado: {excel_path}")


# Carregar dados do Excel para DataFrame
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

# Extrair dados do Excel para um dicionário
data = []
for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row-7, min_col=1, max_col=9, values_only=True):
    if row[0]:  # Verifica se o nome do bairro não é vazio
        data.append({
            'Bairro': row[0],
            'População': row[1],
            'TOTALCasos': row[2],
            'Obitos': row[8]
        })

# Mapear nomes dos bairros para correspondência com JSON
for item in data:
    item['Bairro'] = item['Bairro'].upper()
    item['Bairro'] = {
        'BOM SUCESSO': 'BONSUCESSO',
        'LUCIANO CAVALCANTE': 'ENGENHEIRO LUCIANO CAVALCANTE',
        'PALMEIRAS': 'CONJUNTO PALMEIRAS',
        'PARQUE GENIBAU': 'GENIBAU',
        'PLANALTO AIRTON SENNA': 'PLANALTO AYRTON SENNA',
        'VILA ELLERY': 'ELLERY',
        'VILA MANOEL SATIRO': 'MANOEL SATIRO',
        'SAPIRANGA  COITE': 'SAPIRANGA / COITÉ',
        'BOA VISTA': 'BOA VISTA  CASTELAO',
        'PRAIA DO MEIRELES':'MEIRELES',
        'PAN AMERICANO':'PANAMERICANO',
        'SAPIRANGA COITE':'SAPIRANGA  COITE'
    }.get(item['Bairro'], item['Bairro'])

# Carregar JSON de bairros
with open(r"C:\Users\isabe\OneDrive\Área de Trabalho\NotificaFortalezaTestes\Backend\backendmapadengue\information\arquivos\bairros_latlon.json", 'r', encoding='utf-8') as file:
    bairros_data = json.load(file)

# Criar dicionário com base no JSON
bairros_dict = {feature['properties']['nome']: feature for feature in bairros_data['features']}

# Listar coordenadas dos bairros
coordenadas_bairros = []
for nome_bairro, feature in bairros_dict.items():
    nome_limpo = limpar_nome_bairro(nome_bairro)
    coordenadas = feature['geometry']['coordinates'][0]
    coordenadas_bairros.append({'nome_bairro': nome_limpo, 'coordenadas': coordenadas})


# Endpoint para enviar dados finais como JSON para frontend
@csrf_exempt
def enviar_dataframe_json(request):
    global data
    global coordenadas_bairros

    dados_finais = {}

    for bairro in coordenadas_bairros:
        nome_bairro = bairro['nome_bairro']
        coordenadas = bairro['coordenadas']

        filtro_bairro = next((item for item in data if item['Bairro'] == nome_bairro), None)

        if filtro_bairro:
            dados_bairro = {
                'Bairro': nome_bairro,
                'População': filtro_bairro['População'],
                'TOTALCasos': filtro_bairro['TOTALCasos'],
                'Obitos': filtro_bairro['Obitos'],
                'Coordenadas': coordenadas
            }
        else:
            dados_bairro = {
                'Bairro': nome_bairro,
                'População': None,
                'TOTALCasos': None,
                'Obitos': None,
                'Coordenadas': coordenadas
            }

        dados_finais[nome_bairro] = dados_bairro

    return JsonResponse(list(dados_finais.values()), safe=False)
